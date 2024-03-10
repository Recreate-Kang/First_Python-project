import OpenDartReader
import openpyxl
import os
from openpyxl import load_workbook
import pandas as pd
import csv
import time
api_key= 'e70890fdd042d46e3926b3885ecb7b3f18f5c7a4'

dart = OpenDartReader(api_key)

#print(dart.list)



##종목명및 종목코드 메모장 데이터 정렬 및 가공
Pfile_list = os.listdir('./DataCrawl/Stocks/Kosdaq')
NList=[]
Pfile_name = []
for Pfile in Pfile_list:
    if Pfile.count(".") == 1: 
        name = Pfile.split('.')[0]
        Pfile_name.append(name)
    else:
        for k in range(len(Pfile)-1,0,-1):
            if Pfile[k]=='.':
                Pfile_name.append(Pfile[:k])
                break
print(Pfile_name)

for i in Pfile_name: ##다트에서 종목코드로 제무제표 가져오기
    wb = load_workbook(filename='./DataCrawl/Stocks/Kosdaq/'+i+'.xlsx')
    #print(df)
    print(wb.sheetnames)
    for b in wb.sheetnames:
        if b == 'FSS':
            wb.remove_sheet(wb['FSS'])
            print(wb.sheetnames)
            wb.save(filename='./DataCrawl/Stocks/Kosdaq/'+i+'.xlsx')
            wb = load_workbook(filename='./DataCrawl/Stocks/Kosdaq/'+i+'.xlsx')
    print(wb.sheetnames)
    with pd.ExcelWriter('./DataCrawl/Stocks/Kosdaq/'+i+'.xlsx', mode="a", engine="openpyxl") as writer:
        A=dart.finstate_all(i[:6], 2020)
        print(i[:6])
        print(type(A))
        print(A)
        if A is not None:
            A.to_excel(writer, sheet_name="FSS")
        if A is None:
            NList.append(i)
    time.sleep(0.1)

F= open("D:\Coding/DataCrawl/Stocks/daqNoneList.txt",'w')
for i in NList:
   F.write(i+'\n')
F.close()
'''
A=dart.finstate_all('000020', 2020)
print(type(A))
##wb = load_workbook(filename='./DataCrawl/Stocks/Kospi/000020 동화약품.xlsx')
#ws = wb['FSS']
#ws = A
#wb.save(filename='./DataCrawl/Stocks/Kospi/000020 동화약품.xlsx')
#A.to_excel('./DataCrawl/Stocks/Kospi/000020 동화약품.xlsx', sheet_name='FSS')
#print(type(A))
'''